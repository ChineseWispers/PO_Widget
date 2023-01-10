from tkinter import *
from tkinter import ttk
import openpyxl
import pyautogui as auto

try:
    wb = openpyxl.load_workbook("Processed_POs.xlsx")
    ws = wb["Sheet1"]

    def widget():

        toplevel = Toplevel(root)
        toplevel.title("Unleashed-Lightyear PO Manager")
        toplevel.geometry("405x438+1510+420")
        toplevel.attributes('-topmost', True)
        clipboard = Tk()
        clipboard.withdraw()

        topframe = Frame(toplevel, borderwidth=3, relief="groove")
        topframe.grid(row=0, column=1, rowspan=2, columnspan=6)
        bottomframe = Frame(toplevel, borderwidth=3, relief="groove")
        bottomframe.grid(row=6, column=1, rowspan=2, pady=4)

        def values_not_found():

            label4.config(text="Lightyear not found")
            label6.config(text="Lightyear not found")
            label8.config(text="Lightyear not found")

        def post(*self):

            po_number = string_var2.get()
            search(po_number)

            if string_var3.get():
                invoice_number = string_var3.get()
            else:
                invoice_no_location = auto.locateOnScreen('Invoice.JPG', grayscale=True, confidence=.5)
                if not invoice_no_location:
                    values_not_found()
                    return
                centred_xy = auto.center(invoice_no_location)
                xy = (centred_xy[0] - 60, centred_xy[1] + 30)

                auto.click(xy, clicks=2)
                auto.dragTo(xy[0] - 500, xy[1], button='left')
                auto.hotkey('command', 'c')
                invoice_number = clipboard.clipboard_get()
            if string_var4.get():
                total = string_var4.get()
            else:
                location = auto.locateOnScreen('Total.JPG', grayscale=True, confidence=.5)
                auto.moveTo(location[0], location[1] + 70)
                auto.dragTo(location[0] - 110, location[1] + 70, button='left')
                auto.hotkey('command', 'c')
                total = clipboard.clipboard_get()[1:]
            comments = string_var5.get()

            ws.insert_rows(2)
            ws['A2'] = po_number
            ws['A2'].number_format = '0'
            ws['B2'] = invoice_number
            ws['B2'].number_format = '0'
            ws['C2'] = float(total)
            ws['C2'].number_format = '$#,##0.00'
            ws['D2'] = comments

            label4.config(text=po_number)
            label6.config(text=invoice_number)
            label8.config(text='$'+total)
            label10.config(text=comments)
            post_input.delete(0, 25)
            invoice_input.delete(0, 25)
            total_input.delete(0, 25)

            global b_undo
            b_undo = Button(bottomframe, text="Undo", command=undo, width=6)
            b_undo.grid(row=8, column=3)

            post_input.focus()

            try:
                auto.moveTo(location[0] + 15, location[1] - 53)
            except NameError:
                location = auto.locateOnScreen('Total.JPG', grayscale=True, confidence=.5)
                if location:
                    auto.moveTo(location[0] + 15, location[1] - 53)

        def cancel():
            clipboard.destroy()
            toplevel.destroy()
            root.destroy()

        def search(*from_post):

            for item in listbox.get_children(''):
                listbox.delete(item)

            if len(from_post) != 0:
                if isinstance(from_post[0], str):
                    text = from_post[0]
                else:
                    text = string_var1.get()
            else:
                text = string_var1.get()

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row+1, min_col=1, max_col=1):
                for cell in row:
                    if cell.value:
                        if str(cell.value)[:4] == text[:4]:
                            po_result = cell.value
                            invoice_result = cell.offset(column=1).value
                            total_result = cell.offset(column=2).value
                            listbox.insert('', '0', text=po_result, values=(po_result, invoice_result, total_result))

            ws.delete_rows(ws.max_row)

        def undo():

            ws.delete_rows(2)

            label4.config(text="Undone")
            label6.config(text="Undone")
            label8.config(text="Undone")

            global b_undo
            b_undo.grid_forget()

        label1 = Label(topframe, text="Search PO :     ")
        label1.grid(row=0, column=0)

        string_var1 = StringVar(topframe)
        search_input = Entry(topframe, textvariable=string_var1)
        search_input.grid(row=0, column=1)
        search_input.bind('<Return>', search)
        search_input.bind('<KP_Enter>', search)

        b_search = Button(topframe, text="Search", command=search, width=5)
        b_search.grid(row=0, column=3, padx=12, pady=6)

        treeview_columns = ('PO', 'Invoice', 'Total')
        listbox = ttk.Treeview(topframe, columns=treeview_columns, show='headings')
        listbox.heading('PO', text='PO')
        listbox.heading('Invoice', text='Invoice')
        listbox.heading('Total', text='Total')
        listbox.column('PO', width=6)
        listbox.column('Invoice', width=10)
        listbox.column('Total', width=10)
        listbox.grid(row=1, column=0, rowspan=5, columnspan=4, pady=6, padx=6, sticky='nsew')

        scrollbar = ttk.Scrollbar(topframe, orient='vertical')
        scrollbar.grid(row=1, column=6)
        listbox.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=listbox.yview)

        label3 = Label(bottomframe, text="Post PO :  ")
        label3.grid(row=6, column=0)

        string_var2 = StringVar(bottomframe)
        post_input = Entry(bottomframe, textvariable=string_var2, width=7)
        post_input.grid(row=6, column=1, padx=4)
        post_input.bind('<Return>', post)
        post_input.bind('<KP_Enter>', post)

        label4 = Label(bottomframe, text="", width=14, background="White")
        label4.grid(row=6, column=2)

        b_post = Button(bottomframe, text="Post", command=post, width=6)
        b_post.grid(row=6, column=3, padx=14, pady=12)

        label5 = Label(bottomframe, text="Invoice :  ")
        label5.grid(row=7, column=0)

        string_var3 = StringVar(bottomframe)
        invoice_input = Entry(bottomframe, textvariable=string_var3, width=7)
        invoice_input.grid(row=7, column=1, padx=4)
        invoice_input.bind('<Return>', post)
        invoice_input.bind('<KP_Enter>', post)

        label6 = Label(bottomframe, text="", width=14, background="White")
        label6.grid(row=7, column=2)

        label7 = Label(bottomframe, text="Total :    ")
        label7.grid(row=8, column=0)

        string_var4 = StringVar(bottomframe)
        total_input = Entry(bottomframe, textvariable=string_var4, width=7)
        total_input.grid(row=8, column=1, padx=4, pady=12)
        total_input.bind('<Return>', post)
        total_input.bind('<KP_Enter>', post)

        label8 = Label(bottomframe, text="", width=14, background="White")
        label8.grid(row=8, column=2)

        label9 = Label(bottomframe, text="Comments*: ", font=("Segoe UI", 12))
        label9.grid(row=9, column=0)

        string_var5 = StringVar(bottomframe)
        comment_input = Entry(bottomframe, textvariable=string_var5, width=7)
        comment_input.grid(row=9, column=1, padx=4)
        comment_input.bind('<Return>', post)
        comment_input.bind('<KP_Enter>', post)

        label10 = Label(bottomframe, text="", width=14, background="White")
        label10.grid(row=9, column=2)

        post_input.focus()

        toplevel.protocol("WM_DELETE_WINDOW", cancel)

    root = Tk()
    root.geometry("1x1")
    root.title("Main Window")
    widget()
    root.mainloop()

finally:
    wb.save(filename="Processed_POs.xlsx")
    wb.close()
